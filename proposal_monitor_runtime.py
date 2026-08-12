"""Collect keyword-matched proposal opportunities into an Excel workplan."""

from __future__ import annotations

import argparse
import hashlib
from html import escape
import json
import logging
import os
import re
import smtplib
import time
from datetime import date, datetime, timedelta, timezone
from email.message import EmailMessage
from io import BytesIO
from pathlib import Path
from urllib.parse import urljoin, urlparse, urlunparse
from urllib.robotparser import RobotFileParser

import requests
from bs4 import BeautifulSoup
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from pypdf import PdfReader

# Some public tender portals host malformed but still readable PDFs. pypdf logs
# a warning for every broken cross-reference it repairs, which can overwhelm a
# normal monitoring run. The extraction result remains available to the code.
logging.getLogger("pypdf").setLevel(logging.ERROR)

ROOT = Path(__file__).resolve().parent
OUTPUT = ROOT / "proposal_output"
WORKFLOW_STATE_FILE = OUTPUT / "proposal_workflow_state.json"
PENDING_SOURCE_FILE = ROOT / "Migrations" / "pending_source_intake.json"
DISCOVERY_TERMS = ("tender", "proposal", "rfp", "consultancy", "procurement", "grant", "funding", "certification", "certification offer", "paper proposal", "call for papers")
ICT_TERMS = ("ict", "information technology", "technology", "digital", "software", "computer", "data", "analytics", "artificial intelligence", "machine learning", "cyber", "cloud", "network", "blockchain", "automation", "systems development")
REQUEST_INTERVAL_SECONDS = 1.5
MAX_FETCH_ATTEMPTS = 3
_robots_cache: dict[str, RobotFileParser | None] = {}
_next_request_at: dict[str, float] = {}
DATE_PATTERNS = (
    re.compile(
        r"(?:closing date|deadline|due date|submission date|submit(?:ted)? by|closes? on|on or before|received on or before)"
        r"[^.\n]{0,100}?(\d{1,2}(?:st|nd|rd|th)?\s+(?:January|February|March|April|May|June|July|August|September|October|November|December)\s*,?\s*20\d{2})",
        re.IGNORECASE,
    ),
    re.compile(
        r"(?:closing date|deadline|due date|submission date|submit(?:ted)? by|closes? on|on or before|received on or before)"
        r"[^.\n]{0,100}?((?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{1,2}(?:st|nd|rd|th)?\s*,?\s*20\d{2})",
        re.IGNORECASE,
    ),
)


def atomic_json_write(path: Path, payload: object) -> None:
    """Replace JSON snapshots atomically so readers never see a partial file."""
    path.parent.mkdir(exist_ok=True)
    temporary = path.with_suffix(f"{path.suffix}.tmp")
    temporary.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    temporary.replace(path)


def clean(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def preferred_title(label: str, extracted_text: str) -> str:
    """Replace generic index-link labels with useful text extracted from the document."""
    if label and label.casefold() not in {"download", "loading...", "view", "read more", "click here"}:
        return label
    return clean(extracted_text)[:160] or label or "Untitled proposal"


def canonical_url(url: str) -> str:
    """Make equivalent portal links comparable and remove page fragments."""
    parsed = urlparse(url)
    path = parsed.path.rstrip("/") or "/"
    return urlunparse((parsed.scheme.lower(), parsed.netloc.lower(), path, "", parsed.query, ""))


def allowed(url: str, domains: list[str]) -> bool:
    host = (urlparse(url).hostname or "").lower()
    return any(host == domain or host.endswith(f".{domain}") for domain in domains)


def due_date(text: str) -> str:
    for pattern in DATE_PATTERNS:
        match = pattern.search(text)
        if not match:
            continue
        raw = re.sub(r"(?<=\d)(?:st|nd|rd|th)\b", "", match.group(1), flags=re.IGNORECASE).replace(",", "")
        for format_string in ("%d %B %Y", "%B %d %Y"):
            try:
                return datetime.strptime(raw, format_string).date().isoformat()
            except ValueError:
                pass
    return "Not stated"


def category(matches: list[str]) -> str:
    terms = " ".join(matches).lower()
    if any(term in terms for term in ("digital health africa grant", "ai healthcare africa funding", "telemedicine africa grant", "climate technology africa funding", "smart cities africa grant")):
        return "Digital Health & Climate Tech"
    if any(term in terms for term in ("youth employment africa grant", "youth digital skills africa funding", "women in technology africa funding", "women in stem africa grant", "gender inclusion digital africa grant")):
        return "Youth, Women & Inclusion"
    if "paper proposal" in terms or "call for papers" in terms:
        return "Paper proposal"
    if "grant application" in terms or "funding opportunity" in terms:
        return "Grant"
    if "certification offer" in terms or "certification application" in terms or "apply for certification" in terms:
        return "Certification"
    if any(term in terms for term in ("data science", "data scientist", "data-scientist", "machine learning", "artificial intelligence")):
        return "Data science"
    if any(term in terms for term in ("analytics", "analysis", "analytical", "data analytics", "data analysis")):
        return "Analytics"
    if "training" in terms:
        return "Training"
    return "Other"


def is_ict_related(*values: str) -> bool:
    """Require an ICT signal before accepting broad opportunity types."""
    text = " ".join(str(value or "") for value in values).casefold()
    return any(term in text for term in ICT_TERMS)


def _origin(url: str) -> str:
    parsed = urlparse(url)
    return f"{parsed.scheme}://{parsed.netloc}"


def _request_headers() -> dict[str, str]:
    return {"User-Agent": os.getenv("SCRAPER_USER_AGENT", "ProposalMonitor/1.0 (+contact your technical owner)")}


def _robots_policy(session: requests.Session, url: str) -> RobotFileParser | None:
    """Return a site's robots policy when it is available; cache it per origin."""
    origin = _origin(url)
    if origin in _robots_cache:
        return _robots_cache[origin]
    robots_url = f"{origin}/robots.txt"
    try:
        response = session.get(robots_url, timeout=15, headers=_request_headers())
        if not response.ok:
            _robots_cache[origin] = None
            return None
        policy = RobotFileParser()
        policy.set_url(robots_url)
        policy.parse(response.text.splitlines())
        _robots_cache[origin] = policy
        return policy
    except requests.RequestException:
        # An unavailable robots file is not treated as permission to bypass an
        # access control; the actual request will still respect HTTP responses.
        _robots_cache[origin] = None
        return None


def _pace(origin: str) -> None:
    interval = max(0.5, float(os.getenv("SCRAPER_REQUEST_INTERVAL_SECONDS", str(REQUEST_INTERVAL_SECONDS))))
    wait = _next_request_at.get(origin, 0) - time.monotonic()
    if wait > 0:
        time.sleep(wait)
    _next_request_at[origin] = time.monotonic() + interval


def fetch(session: requests.Session, url: str) -> tuple[str, str]:
    """Fetch a public page politely; do not bypass robots rules or access controls."""
    origin = _origin(url)
    policy = _robots_policy(session, url)
    user_agent = _request_headers()["User-Agent"]
    if policy and not policy.can_fetch(user_agent, url):
        raise requests.exceptions.RequestException(f"robots.txt disallows this crawler: {url}")

    response = None
    for attempt in range(MAX_FETCH_ATTEMPTS):
        _pace(origin)
        response = session.get(url, timeout=30, headers=_request_headers())
        if response.status_code not in (429, 500, 502, 503, 504) or attempt == MAX_FETCH_ATTEMPTS - 1:
            break
        retry_after = response.headers.get("Retry-After")
        try:
            delay = min(60.0, max(1.0, float(retry_after))) if retry_after else 2.0 * (attempt + 1)
        except ValueError:
            delay = 2.0 * (attempt + 1)
        print(f"Temporary response {response.status_code} from {origin}; waiting {delay:g}s before retrying.")
        time.sleep(delay)
    assert response is not None
    response.raise_for_status()
    content_type = response.headers.get("content-type", "").lower()
    if "pdf" in content_type or urlparse(url).path.lower().endswith(".pdf"):
        text = "\n".join(page.extract_text() or "" for page in PdfReader(BytesIO(response.content)).pages)
        return "pdf", text
    return "html", response.text


def collect(session: requests.Session, source: dict, keywords: list[str]) -> list[dict[str, str]]:
    candidates: dict[str, str] = {}
    start_pages = {canonical_url(url) for url in source["start_urls"]}
    for start_url in source["start_urls"]:
        try:
            _, body = fetch(session, start_url)
            soup = BeautifulSoup(body, "html.parser")
        except requests.RequestException as error:
            print(f"Warning: could not read {start_url}: {error}")
            continue
        for anchor in soup.find_all("a", href=True):
            link = urljoin(start_url, anchor["href"])
            canonical = canonical_url(link)
            label = clean(anchor.get_text(" ", strip=True))
            # Navigation links such as "All Tenders" are portal indexes, not
            # opportunities. They otherwise match generic keywords in their page text.
            is_portal_index = label.casefold() in {"all tenders", "tenders", "procurement"}
            if allowed(canonical, source["allowed_domains"]) and canonical not in start_pages and not is_portal_index:
                candidates[canonical] = label

    results = []
    for link, label in list(candidates.items())[: source.get("max_links", 80)]:
        hint = f"{label} {link}".lower()
        if not any(word.lower() in hint for word in keywords) and not any(term in hint for term in DISCOVERY_TERMS):
            continue
        try:
            kind, body = fetch(session, link)
        except (requests.RequestException, OSError):
            continue
        if kind == "pdf":
            text = clean(body)
            title = preferred_title(label, text)
        else:
            soup = BeautifulSoup(body, "html.parser")
            title_tag = soup.find("h1") or soup.find("title")
            text = clean(soup.get_text(" ", strip=True))
            title = preferred_title(clean(title_tag.get_text(" ", strip=True) if title_tag else label), text)
        matches = [word for word in keywords if word.lower() in f"{title} {text}".lower()]
        if matches and is_ict_related(title, text):
            results.append({"name": title or label or "Untitled proposal", "category": category(matches), "link": link, "due_date": due_date(text), "source": source["name"], "keywords": ", ".join(matches)})
    print(f"Checked {source['name']}: {len(candidates)} candidate link(s), {len(results)} matching proposal(s).")
    return results


def active_sources(config: dict) -> list[dict]:
    """Validate source entries and return only sites approved for monitoring."""
    valid = []
    for source in config.get("sources", []):
        if not source.get("active", True):
            print(f"Skipping inactive source: {source.get('name', 'Unnamed source')}")
            continue
        required = ("name", "start_urls", "allowed_domains")
        if not all(source.get(field) for field in required):
            print(f"Skipping incomplete source: {source.get('name', 'Unnamed source')}")
            continue
        valid.append(source)
    return valid


def format_sheet(sheet, headings: list[str]) -> None:
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index in range(1, len(headings) + 1):
        sheet.column_dimensions[get_column_letter(index)].width = min(60, max(14, max(len(str(row[index - 1].value or "")) for row in sheet.iter_rows()) + 2))


def add_sheet(book: Workbook, name: str, headings: list[str], rows: list[list[str]]) -> None:
    sheet = book.create_sheet(name)
    sheet.append(headings)
    for row in rows:
        sheet.append(row)
    format_sheet(sheet, headings)


def item_id(item: dict[str, str]) -> str:
    return hashlib.sha256(f"{item['link']}|{item['name']}".encode()).hexdigest()[:16]


def load_workflow_state() -> dict[str, dict[str, str]]:
    if not WORKFLOW_STATE_FILE.exists():
        return {}
    return json.loads(WORKFLOW_STATE_FILE.read_text(encoding="utf-8"))


def review_due_date(item: dict[str, str], today: date) -> str:
    """Review deadline is three days before tender close, or two days after discovery."""
    if item["due_date"] != "Not stated":
        return max(today, datetime.fromisoformat(item["due_date"]).date() - timedelta(days=3)).isoformat()
    return (today + timedelta(days=2)).isoformat()


def is_expired(item: dict[str, str], today: date | None = None) -> bool:
    if item.get("due_date") in (None, "", "Not stated"):
        return False
    return datetime.fromisoformat(item["due_date"]).date() < (today or date.today())


def merge_with_dashboard_snapshot(fresh: list[dict[str, str]]) -> list[dict[str, str]]:
    """Keep previously found, still-active opportunities between monitor runs.

    A fresh record with the same official link replaces the old copy; expired
    records are removed. This prevents a temporary source-page change from
    making valid earlier proposals disappear from the dashboard or workbook.
    """
    snapshot = OUTPUT / "proposals.json"
    existing: list[dict[str, str]] = []
    if snapshot.exists():
        try:
            existing = json.loads(snapshot.read_text(encoding="utf-8")).get("proposals", [])
        except (json.JSONDecodeError, OSError):
            existing = []
    merged = {item.get("link"): item for item in [*existing, *fresh] if item.get("link")}
    expired = [item for item in merged.values() if is_expired(item)]
    if expired:
        archive_file = OUTPUT / "previous_opportunities.json"
        try:
            archived = json.loads(archive_file.read_text(encoding="utf-8")).get("proposals", []) if archive_file.exists() else []
        except (json.JSONDecodeError, OSError):
            archived = []
        now = datetime.now(timezone.utc).isoformat()
        archive = {item.get("link"): item for item in archived if item.get("link")}
        for item in expired:
            archive[item["link"]] = {**item, "archived_at": now}
        atomic_json_write(archive_file, {"generated_at": now, "count": len(archive), "proposals": list(archive.values())})
    return [item for item in merged.values() if not is_expired(item)]


def sync_workflow(proposals: list[dict[str, str]]) -> dict[str, dict[str, str]]:
    """Create persistent review tasks for newly discovered proposals."""
    state = load_workflow_state()
    today = date.today()
    for item in proposals:
        proposal_key = item_id(item)
        task = state.setdefault(proposal_key, {
            "status": "Review required",
            "owner": "Project team",
            "first_seen": today.isoformat(),
            "review_due_date": review_due_date(item, today),
        })
        task["title"] = item["name"]
        task["link"] = item["link"]
        task["last_seen"] = today.isoformat()
    atomic_json_write(WORKFLOW_STATE_FILE, state)
    return state


def source_register_rows(sources: list[dict] | None) -> list[list[str]]:
    """Show both configured and user-submitted sources in the workbook."""
    rows = []
    for source in sources or []:
        rows.append([source.get("name", ""), ", ".join(source.get("start_urls", [])), source.get("source_type", ""), "", source.get("check_frequency", ""), str(source.get("active", True)).lower(), source.get("notes", "")])
    if not PENDING_SOURCE_FILE.exists():
        return rows
    try:
        pending = json.loads(PENDING_SOURCE_FILE.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        print("Warning: pending source intake is not valid JSON; it was omitted from the workbook.")
        return rows
    for item in pending:
        rows.append(["User submitted", item.get("url", ""), "Pending review", "", "", "false", f"Submitted {item.get('submitted_at', '')}"])
    return rows


def export(proposals: list[dict[str, str]], output: Path, workflow_state: dict[str, dict[str, str]], sources: list[dict] | None = None, keywords: list[str] | None = None) -> None:
    book = Workbook()
    proposals_sheet = book.active
    proposals_sheet.title = "Proposals"
    headings = ["Tracker ID", "Proposal name", "Category", "Website link", "Due date", "Source", "Matched keywords", "Relevance", "Why it matches", "Eligibility notes", "Recommended action", "Status", "Owner", "Date found", "Notes"]
    proposals_sheet.append(headings)
    for item in sorted(proposals, key=lambda row: (row["due_date"] == "Not stated", row["due_date"], row["name"].lower())):
        deadline = datetime.fromisoformat(item["due_date"]).date() if item["due_date"] != "Not stated" else "Not stated"
        task = workflow_state[item_id(item)]
        proposals_sheet.append([item_id(item), item["name"], item["category"], item["link"], deadline, item["source"], item["keywords"], item.get("relevance_score", ""), item.get("match_reason", ""), item.get("eligibility_notes", ""), item.get("recommended_action", ""), task["status"], task["owner"], date.today(), ""])
    format_sheet(proposals_sheet, headings)
    for cell in proposals_sheet["E"][1:]:
        if isinstance(cell.value, date):
            cell.number_format = "yyyy-mm-dd"
    for cell in proposals_sheet["N"][1:]:
        cell.number_format = "yyyy-mm-dd"
    run_date = date.today()
    workplan_headings = ["Task", "Owner", "Start date", "Due date", "Status", "Dependencies", "Notes"]
    workplan_rows = [
        ["Provide and prioritise source website links", "Olivia", run_date, run_date + timedelta(days=2), "Not started", "", "Add approved ministry and donor portals to the source intake."],
        ["Confirm keywords and exclusion terms", "Olivia", run_date, run_date + timedelta(days=2), "Not started", "Source list", "Confirm analytics, data science, training, and exclusion terms."],
        ["Add approved sources and complete a test scrape", "Technical owner", run_date + timedelta(days=3), run_date + timedelta(days=5), "Not started", "Olivia source links", "Activate and validate each source before recurring monitoring."],
        ["Configure Tuesday/Thursday myGov alert recipient", "Olivia / Technical owner", run_date + timedelta(days=3), run_date + timedelta(days=7), "Not started", "SMTP access", "Set alert recipient and SMTP details in .env."],
        ["Review new opportunities and assign a proposal owner", "Project team", run_date + timedelta(days=5), run_date + timedelta(days=12), "Ongoing", "Proposal tracker", "Update proposal status, owner, and notes after each scan."],
        ["Define content-folder approval and posting workflow", "Olivia / Content owner", run_date + timedelta(days=7), run_date + timedelta(days=14), "Not started", "Content folder", "Confirm approval steps before connecting posting automation."],
        ["Review performance and refine keywords", "Project team", run_date + timedelta(days=14), run_date + timedelta(days=21), "Not started", "Two weeks of results", "Remove irrelevant results and add useful new terms."],
    ]
    add_sheet(book, "Workplan", workplan_headings, workplan_rows)
    for cell in book["Workplan"]["C"][1:] + book["Workplan"]["D"][1:]:
        cell.number_format = "yyyy-mm-dd"
    add_sheet(book, "Website Register", ["Source name", "Website URL", "Type", "Priority", "Check frequency", "Active", "Notes"], source_register_rows(sources))
    add_sheet(book, "Keywords", ["Keyword", "Category", "Include / Exclude", "Notes"], [[keyword, category([keyword]), "Include", ""] for keyword in (keywords or [])])
    add_sheet(book, "Content Calendar", ["Content item", "Platform", "Owner", "Approval status", "Posting date", "Published", "Notes"], [])
    queue_rows = []
    for item in proposals:
        task = workflow_state[item_id(item)]
        queue_rows.append([item_id(item), item["name"], "Review proposal and assign a bid decision", task["owner"], datetime.fromisoformat(task["review_due_date"]).date(), task["status"], item["link"]])
    add_sheet(book, "Workflow Queue", ["Tracker ID", "Proposal", "Next action", "Owner", "Review due date", "Status", "Link"], queue_rows)
    for cell in book["Workflow Queue"]["E"][1:]:
        cell.number_format = "yyyy-mm-dd"
    add_sheet(book, "Workflows", ["Workflow", "Trigger", "Automated action", "Owner", "Output / escalation"], [["Proposal discovery", "Scheduled monitoring finds a keyword match", "Create or update a persistent review task", "Project team", "Workflow Queue; review due date is three days before tender close."], ["myGov alert", "Tuesday/Thursday scheduler finds new matches", "Send one email per new opportunity", "Project team", "Scheduler waits for SMTP configuration and prevents repeat alerts."], ["Proposal review", "Workflow Queue contains Review required", "Assess relevance and set bid decision", "Proposal owner", "Update the workflow state using the command below."], ["Content publishing", "Content receives approval", "Queue item for posting automation", "Content owner", "To be connected when the content folder/platform is provided."]])
    output.parent.mkdir(exist_ok=True)
    book.save(output)


def write_dashboard_results(proposals: list[dict[str, str]], workflow_state: dict[str, dict[str, str]]) -> None:
    """Write the compact result snapshot consumed by the web dashboard."""
    rows = []
    for proposal in proposals:
        task = workflow_state.get(item_id(proposal), {})
        rows.append({
            "id": item_id(proposal),
            "name": proposal.get("name", "Untitled proposal"),
            "category": proposal.get("category", "Other"),
            "link": proposal.get("link", ""),
            "due_date": proposal.get("due_date", "Not stated"),
            "source": proposal.get("source", "Unknown source"),
            "keywords": proposal.get("keywords", ""),
            "relevance_score": proposal.get("relevance_score", ""),
            "match_reason": proposal.get("match_reason", ""),
            "eligibility_notes": proposal.get("eligibility_notes", ""),
            "recommended_action": proposal.get("recommended_action", "Review"),
            "status": task.get("status", "Review required"),
            "owner": task.get("owner", "Project team"),
            "review_due_date": task.get("review_due_date", ""),
        })
    atomic_json_write(OUTPUT / "proposals.json", {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "count": len(rows),
        "proposals": rows,
    })


def proposal_alert_text(items: list[dict[str, str]]) -> str:
    """Build the plain-text fallback for a scheduled priority alert."""
    return "High- and medium-priority proposals ready for review:\n\n" + "\n\n".join(
        f"{item['name']}\n"
        f"Category: {item['category']}\n"
        f"Due: {item['due_date']}\n"
        f"View opportunity: {item['link']}"
        for item in items
    )


def proposal_alert_html(items: list[dict[str, str]]) -> str:
    """Build a self-contained, email-client-safe HTML myGov alert."""
    count = len(items)
    cards = "".join(
        f"""
        <tr>
          <td style="padding:0 28px 18px;">
            <table role="presentation" width="100%" cellspacing="0" cellpadding="0" border="0" style="border:1px solid #dbe5ee;border-radius:10px;background:#ffffff;">
              <tr><td style="padding:20px 20px 8px;font-family:Arial,sans-serif;color:#172b4d;font-size:18px;font-weight:700;line-height:26px;">{escape(item['name'])}</td></tr>
              <tr><td style="padding:0 20px 8px;font-family:Arial,sans-serif;color:#52657a;font-size:14px;line-height:21px;"><strong style="color:#172b4d;">Category:</strong> {escape(item['category'])}&nbsp;&nbsp; <strong style="color:#172b4d;">Due:</strong> {escape(item['due_date'])}</td></tr>
              <tr><td style="padding:6px 20px 20px;"><a href="{escape(item['link'], quote=True)}" style="display:inline-block;background:#0f766e;border-radius:6px;color:#ffffff;font-family:Arial,sans-serif;font-size:14px;font-weight:700;line-height:18px;padding:11px 16px;text-decoration:none;">View opportunity</a></td></tr>
            </table>
          </td>
        </tr>"""
        for item in items
    )
    label = "opportunity" if count == 1 else "opportunities"
    return f"""<!doctype html>
<html lang="en"><body style="margin:0;padding:0;background:#f4f7fb;">
  <table role="presentation" width="100%" cellspacing="0" cellpadding="0" border="0" style="background:#f4f7fb;"><tr><td align="center" style="padding:32px 12px;">
    <table role="presentation" width="100%" cellspacing="0" cellpadding="0" border="0" style="max-width:640px;background:#ffffff;border-radius:14px;overflow:hidden;">
      <tr><td style="padding:28px;background:#102b46;font-family:Arial,sans-serif;color:#ffffff;"><div style="font-size:12px;font-weight:700;letter-spacing:1.2px;text-transform:uppercase;color:#71dac2;">Proposal Monitor</div><div style="padding-top:9px;font-size:26px;font-weight:700;line-height:34px;">Priority {label} for review</div><div style="padding-top:8px;font-size:15px;line-height:22px;color:#c9d7e4;">{count} high- or medium-priority match{'es' if count != 1 else ''} in this scheduled briefing.</div></td></tr>
      <tr><td style="padding-top:24px;">{cards}</td></tr>
      <tr><td style="padding:20px 28px 28px;font-family:Arial,sans-serif;color:#6b7b8f;font-size:12px;line-height:18px;">This automated alert was sent by Proposal Monitor. Review each opportunity and its requirements before taking action.</td></tr>
    </table>
  </td></tr></table>
</body></html>"""


def email_alert(items: list[dict[str, str]]) -> bool:
    required = [os.getenv("ALERT_EMAIL_TO"), os.getenv("SMTP_HOST"), os.getenv("SMTP_USERNAME"), os.getenv("SMTP_PASSWORD")]
    if not all(required):
        print("No email sent: set ALERT_EMAIL_TO and SMTP_* in .env.")
        return False
    message = EmailMessage()
    message["Subject"] = f"Proposal Monitor priority briefing: {len(items)} opportunity(s)"
    message["From"] = os.getenv("SMTP_FROM") or os.getenv("SMTP_USERNAME")
    message["To"] = os.environ["ALERT_EMAIL_TO"]
    message.set_content(proposal_alert_text(items))
    message.add_alternative(proposal_alert_html(items), subtype="html")
    with smtplib.SMTP(os.environ["SMTP_HOST"], int(os.getenv("SMTP_PORT", "587"))) as smtp:
        smtp.starttls()
        smtp.login(os.environ["SMTP_USERNAME"], os.environ["SMTP_PASSWORD"])
        smtp.send_message(message)
    return True


def main() -> int:
    parser = argparse.ArgumentParser(description="Collect matching proposals and export a workplan.")
    parser.add_argument("--config", type=Path, default=ROOT / "Migrations" / "proposal_source.json")
    parser.add_argument("--output", type=Path, default=OUTPUT / "proposals.xlsx")
    parser.add_argument("--mygov-alert", action="store_true")
    parser.add_argument("--set-workflow", nargs=3, metavar=("TRACKER_ID", "STATUS", "OWNER"), help="Update a review task, for example: --set-workflow abc123 Reviewing Olivia")
    args = parser.parse_args()
    load_dotenv(ROOT / ".env")
    if args.set_workflow:
        tracker_id, status, owner = args.set_workflow
        state = load_workflow_state()
        if tracker_id not in state:
            raise SystemExit(f"Unknown tracker ID: {tracker_id}")
        state[tracker_id]["status"] = status
        state[tracker_id]["owner"] = owner
        WORKFLOW_STATE_FILE.write_text(json.dumps(state, indent=2), encoding="utf-8")
        print(f"Updated {tracker_id}: {status} ({owner})")
        return 0
    config = json.loads(args.config.read_text(encoding="utf-8-sig"))
    sources = active_sources(config)
    if not sources:
        raise SystemExit("No active, valid sources found. Add approved websites to Migrations/proposal_source.json.")
    print(f"Checking {len(sources)} active source(s) for: {', '.join(config['keywords'])}")
    with requests.Session() as session:
        found = [item for source in sources for item in collect(session, source, config["keywords"])]
    proposals = list({item["link"]: item for item in found}.values())
    expired = [item for item in proposals if is_expired(item)]
    proposals = merge_with_dashboard_snapshot(proposals)
    workflow_state = sync_workflow(proposals)
    export(proposals, args.output, workflow_state, config.get("sources"), config.get("keywords"))
    write_dashboard_results(proposals, workflow_state)
    print(f"Saved {len(proposals)} active matching proposal(s) to {args.output}; excluded {len(expired)} expired item(s).")
    if args.mygov_alert:
        state = OUTPUT / "mygov_seen.json"
        seen = set(json.loads(state.read_text()) if state.exists() else [])
        new = [item for item in proposals if item["source"].lower().startswith("mygov") and item_id(item) not in seen]
        if new and email_alert(new):
            state.write_text(json.dumps(sorted(seen | {item_id(item) for item in new}), indent=2))
        elif not new:
            print("No new myGov matches.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
